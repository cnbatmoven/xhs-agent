from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import os
import random
import re
import sys
import time
import urllib.parse
import urllib.error
import urllib.request
import zipfile
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


NS_MAIN = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
NS_REL = {"r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
NS_PKG_REL = {"pr": "http://schemas.openxmlformats.org/package/2006/relationships"}


@dataclass
class SourceNote:
    row_number: int
    title: str = ""
    url: str = ""
    likes: int | None = None
    collects: int | None = None
    comments: int | None = None
    shares: int | None = None
    note_id: str = ""
    xsec_token: str = ""
    xsec_source: str = ""
    author_id: str = ""
    author_url: str = ""
    pgy_url: str = ""
    pgy_price: float | None = None
    pgy_image_price: float | None = None
    pgy_video_price: float | None = None


@dataclass
class NoteResult:
    row_number: int
    title: str
    url: str
    cover: str
    copywriting: str
    topics: str
    author_nickname: str
    author_id: str
    author_url: str
    fans_count: int | None
    top_comments: str
    likes: int | None
    collects: int | None
    comments: int | None
    shares: int | None
    total_interactions: int | None
    pgy_url: str
    pgy_price: float | None
    pgy_image_price: float | None
    pgy_video_price: float | None
    cpe: float | None
    image_cpe: float | None
    video_cpe: float | None
    content_type: str
    content_group: str
    title_pattern: str
    selling_points: str
    engagement_type: str
    creative_advice: str
    audience_strategy: str
    status: str
    llm_status: str = ""
    llm_model: str = ""
    error: str = ""


def safe_print(text: str) -> None:
    try:
        print(text)
    except UnicodeEncodeError:
        sys.stdout.buffer.write((str(text) + "\n").encode("utf-8", errors="replace"))


def col_to_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    value = 0
    for ch in letters:
        value = value * 26 + (ord(ch.upper()) - ord("A") + 1)
    return value - 1


def parse_int(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    multiplier = 1
    if text.endswith("万"):
        multiplier = 10000
        text = text[:-1]
    elif text.lower().endswith("w"):
        multiplier = 10000
        text = text[:-1]
    elif text.lower().endswith("k"):
        multiplier = 1000
        text = text[:-1]
    try:
        return int(float(text) * multiplier)
    except ValueError:
        match = re.search(r"\d+(?:\.\d+)?", text)
        if not match:
            return None
        return int(float(match.group(0)) * multiplier)


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "").replace("¥", "").replace("￥", "")
    multiplier = 1.0
    if text.endswith("万"):
        multiplier = 10000.0
        text = text[:-1]
    elif text.lower().endswith("w"):
        multiplier = 10000.0
        text = text[:-1]
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0)) * multiplier


def first_record_value(record: dict[str, Any], names: list[str]) -> Any:
    for name in names:
        value = record.get(name)
        if value not in (None, ""):
            return value
    return ""


class BrokenDimensionXlsxReader:
    """Reads inline-string xlsx exports even when worksheet dimension is wrong."""

    def __init__(self, path: Path):
        self.path = path

    def read_sheet(self, sheet_name: str) -> list[dict[str, Any]]:
        with zipfile.ZipFile(self.path) as zf:
            target = self._sheet_target(zf, sheet_name)
            shared_strings = self._shared_strings(zf)
            rows = self._parse_sheet(zf, target, shared_strings)
        if not rows:
            return []
        headers = [str(v).strip() if v is not None else "" for v in rows[0]]
        records: list[dict[str, Any]] = []
        for idx, row in enumerate(rows[1:], start=2):
            if not any(cell not in (None, "") for cell in row):
                continue
            record = {"__row_number": idx}
            for col_idx, header in enumerate(headers):
                if not header:
                    continue
                record[header] = row[col_idx] if col_idx < len(row) else ""
            records.append(record)
        return records

    def _sheet_target(self, zf: zipfile.ZipFile, sheet_name: str) -> str:
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels.findall("pr:Relationship", NS_PKG_REL)
        }
        for sheet in workbook.findall(".//m:sheet", NS_MAIN):
            name = sheet.attrib.get("name")
            rel_id = sheet.attrib.get(f"{{{NS_REL['r']}}}id")
            if name == sheet_name and rel_id in rel_map:
                target = rel_map[rel_id].replace("\\", "/")
                target = target.lstrip("/")
                if target.startswith("xl/"):
                    return target
                if target.startswith("worksheets/"):
                    return f"xl/{target}"
                return f"xl/{target}"
        raise ValueError(f"Sheet not found: {sheet_name}")

    def _shared_strings(self, zf: zipfile.ZipFile) -> list[str]:
        if "xl/sharedStrings.xml" not in zf.namelist():
            return []
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
        strings = []
        for si in root.findall("m:si", NS_MAIN):
            strings.append("".join(t.text or "" for t in si.findall(".//m:t", NS_MAIN)))
        return strings

    def _parse_sheet(
        self, zf: zipfile.ZipFile, target: str, shared_strings: list[str]
    ) -> list[list[Any]]:
        root = ET.fromstring(zf.read(target))
        parsed: list[list[Any]] = []
        for row in root.findall(".//m:row", NS_MAIN):
            cells: list[Any] = []
            for cell in row.findall("m:c", NS_MAIN):
                idx = col_to_index(cell.attrib["r"])
                while len(cells) <= idx:
                    cells.append("")
                cells[idx] = self._cell_value(cell, shared_strings)
            parsed.append(cells)
        return parsed

    def _cell_value(self, cell: ET.Element, shared_strings: list[str]) -> Any:
        cell_type = cell.attrib.get("t")
        if cell_type == "inlineStr":
            return "".join(t.text or "" for t in cell.findall(".//m:t", NS_MAIN))
        value = cell.find("m:v", NS_MAIN)
        if value is None or value.text is None:
            return ""
        if cell_type == "s":
            idx = int(value.text)
            return shared_strings[idx] if idx < len(shared_strings) else ""
        if cell_type == "n":
            number = float(value.text)
            return int(number) if number.is_integer() else number
        return value.text


def load_source_notes(input_path: Path) -> list[SourceNote]:
    records = BrokenDimensionXlsxReader(input_path).read_sheet("笔记明细")
    notes: list[SourceNote] = []
    for record in records:
        url = (
            str(record.get("笔记url", ""))
            or str(record.get("笔记链接", ""))
            or str(record.get("链接", ""))
        ).strip()
        if not url:
            continue
        notes.append(
            SourceNote(
                row_number=int(record["__row_number"]),
                title=str(record.get("笔记标题", "")).strip(),
                url=url,
                likes=parse_int(record.get("点赞数")),
                collects=parse_int(record.get("收藏数")),
                comments=parse_int(record.get("评论数")),
                shares=parse_int(record.get("分享数")),
                author_id=str(first_record_value(record, ["达人ID", "达人id", "作者ID", "作者id", "博主ID", "博主id", "user_id"])).strip(),
                author_url=str(first_record_value(record, ["达人链接", "作者链接", "博主链接", "主页链接"])).strip(),
                pgy_url=str(first_record_value(record, ["蒲公英链接", "蒲公英达人链接", "蒲公英后台链接"])).strip(),
                pgy_price=parse_float(first_record_value(record, ["蒲公英报价", "报价", "价格", "合作报价", "图文报价"])),
                pgy_image_price=parse_float(first_record_value(record, ["蒲公英图文报价", "图文报价", "图文笔记一口价"])),
                pgy_video_price=parse_float(first_record_value(record, ["蒲公英视频报价", "视频报价", "视频笔记一口价"])),
                **parse_note_url(url),
            )
        )
    return notes


def parse_note_url(url: str) -> dict[str, str]:
    parsed = urllib.parse.urlparse(url)
    params = urllib.parse.parse_qs(parsed.query)
    return {
        "note_id": parsed.path.rstrip("/").split("/")[-1],
        "xsec_token": params.get("xsec_token", [""])[0],
        "xsec_source": params.get("xsec_source", [""])[0] or "pc_pgy",
    }


class XhsCrawler:
    def __init__(
        self,
        headless: bool = False,
        profile: Path | None = None,
        timeout_ms: int = 30000,
        download_covers: bool = False,
        cover_dir: Path | None = None,
        browser_executable: Path | None = None,
        cdp_url: str | None = None,
        login_first: bool = False,
        crawl_delay: float = 1.5,
        stop_on_rate_limit: bool = True,
        rate_limit_cooldown: int = 0,
        comment_api: bool = True,
    ):
        self.headless = headless
        self.profile = profile
        self.timeout_ms = timeout_ms
        self.download_covers = download_covers
        self.cover_dir = cover_dir
        self.browser_executable = browser_executable or find_chrome_executable()
        self.cdp_url = cdp_url
        self.login_first = login_first
        self.crawl_delay = crawl_delay
        self.stop_on_rate_limit = stop_on_rate_limit
        self.rate_limit_cooldown = rate_limit_cooldown
        self.comment_api = comment_api

    def crawl_many(self, notes: list[SourceNote]) -> list[dict[str, Any]]:
        try:
            return self._crawl_with_playwright(notes)
        except Exception as exc:
            safe_print(f"Playwright unavailable or failed, fallback to static fetch: {exc}")
            return [self._static_fetch(note) for note in notes]

    def _crawl_with_playwright(self, notes: list[SourceNote]) -> list[dict[str, Any]]:
        from playwright.sync_api import sync_playwright

        results: list[dict[str, Any]] = []
        with sync_playwright() as p:
            if self.cdp_url:
                browser = p.chromium.connect_over_cdp(self.cdp_url)
                context = browser.contexts[0] if browser.contexts else browser.new_context()
                page = context.pages[0] if context.pages else context.new_page()
                self._wait_for_manual_login(page)
                for idx, note in enumerate(notes, start=1):
                    safe_print(f"[{idx}/{len(notes)}] crawl {note.url}")
                    result = self._crawl_page(page, note)
                    results.append(result)
                    if self._should_stop_for_rate_limit(result, notes, idx, results):
                        break
                    time.sleep(self.crawl_delay)
                return results

            launch_options: dict[str, Any] = {
                "headless": self.headless,
            }
            if self.browser_executable:
                launch_options["executable_path"] = str(self.browser_executable)
            if self.profile:
                context = p.chromium.launch_persistent_context(
                    str(self.profile),
                    **launch_options,
                    viewport={"width": 1366, "height": 900},
                )
                page = context.pages[0] if context.pages else context.new_page()
            else:
                browser = p.chromium.launch(**launch_options)
                context = browser.new_context(viewport={"width": 1366, "height": 900})
                page = context.new_page()

            self._wait_for_manual_login(page)
            for idx, note in enumerate(notes, start=1):
                safe_print(f"[{idx}/{len(notes)}] crawl {note.url}")
                result = self._crawl_page(page, note)
                results.append(result)
                if self._should_stop_for_rate_limit(result, notes, idx, results):
                    break
                time.sleep(self.crawl_delay)
            context.close()
        return results

    def _should_stop_for_rate_limit(
        self,
        result: dict[str, Any],
        notes: list[SourceNote],
        current_index: int,
        results: list[dict[str, Any]],
    ) -> bool:
        if result.get("status") != "rate_limited":
            return False
        if self.rate_limit_cooldown > 0:
            safe_print(f"Rate limited. Cooling down for {self.rate_limit_cooldown} seconds...")
            time.sleep(self.rate_limit_cooldown)
            return False
        if not self.stop_on_rate_limit:
            return False
        remaining = len(notes) - current_index
        safe_print(f"Rate limited. Stop crawling and mark {remaining} remaining notes as skipped.")
        for _ in notes[current_index:]:
            results.append(
                {
                    "status": "skipped_rate_limit",
                    "error": "Skipped after security verification to avoid worsening rate limit.",
                }
            )
        return True

    def _crawl_page(self, page: Any, note: SourceNote) -> dict[str, Any]:
        api_comments: list[str] = []

        def on_response(response: Any) -> None:
            try:
                if "comment" not in response.url:
                    return
                if response.status != 200:
                    return
                payload = response.json()
                api_comments.extend(extract_comments_from_payload(payload, limit=20))
            except Exception:
                return

        page.on("response", on_response)
        try:
            page.goto(note.url, wait_until="domcontentloaded", timeout=self.timeout_ms)
            page.wait_for_timeout(2500)
            html_text = page.content()
            page_text = safe_page_text(page)
            if is_security_verification(page_text, html_text):
                return {
                    "status": "rate_limited",
                    "error": "Xiaohongshu security verification: please wait and retry with a larger --crawl-delay.",
                }
            if is_missing_note_page(page_text, html_text):
                return {
                    "status": "missing",
                    "title": "小红书 - 你访问的页面不见了",
                    "error": "Xiaohongshu note page is missing or unavailable.",
                }
            data = extract_from_html(html_text, page_text, note.note_id)
            if is_placeholder_copy(data.get("copywriting")) and not data.get("topics"):
                data["copywriting"] = ""
            if not data.get("top_comments"):
                dom_comments = extract_dom_comments(page, limit=20)
                data["top_comments"] = merge_comments(api_comments, dom_comments, limit=20)
            if self.comment_api and count_numbered_lines(data.get("top_comments")) < 20:
                signed_comments, comment_error = fetch_signed_note_comments(page, note, limit=20)
                if signed_comments:
                    data["top_comments"] = merge_comments(
                        signed_comments,
                        str(data.get("top_comments") or ""),
                        limit=20,
                    )
                if comment_error:
                    existing_error = str(data.get("error") or "").strip()
                    data["error"] = "; ".join(
                        part for part in [existing_error, f"comment_api: {comment_error}"] if part
                    )
            if data.get("author_url") and not data.get("fans_count"):
                profile = {}
                if self.comment_api and data.get("author_id"):
                    profile, profile_error = fetch_signed_author_profile(page, str(data["author_id"]))
                    if profile_error:
                        existing_error = str(data.get("error") or "").strip()
                        data["error"] = "; ".join(
                            part for part in [existing_error, f"author_api: {profile_error}"] if part
                        )
                if not profile:
                    profile = fetch_author_profile(page, str(data["author_url"]))
                if profile.get("fans_count") is not None:
                    data["fans_count"] = profile["fans_count"]
                if profile.get("author_nickname") and not data.get("author_nickname"):
                    data["author_nickname"] = profile["author_nickname"]
            data["cover"] = self._maybe_download_cover(data.get("cover"), note.row_number)
            data["status"] = "ok"
            return data
        except Exception as exc:
            return {"status": "failed", "error": str(exc)}
        finally:
            try:
                page.remove_listener("response", on_response)
            except Exception:
                pass

    def _wait_for_manual_login(self, page: Any) -> None:
        if not self.login_first:
            return
        safe_print("Opening Xiaohongshu login page. Scan the QR code, then press Enter here.")
        page.goto("https://www.xiaohongshu.com", wait_until="domcontentloaded", timeout=self.timeout_ms)
        try:
            input("Login finished? Press Enter to start crawling...")
        except EOFError:
            safe_print("No interactive stdin; continue crawling after 60 seconds.")
            page.wait_for_timeout(60000)

    def _static_fetch(self, note: SourceNote) -> dict[str, Any]:
        try:
            headers = {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0 Safari/537.36"
                ),
                "Referer": "https://www.xiaohongshu.com/",
            }
            body, _ = http_get(note.url, headers=headers)
            text = body.decode("utf-8", errors="replace")
            if is_security_verification(text, text):
                return {
                    "status": "rate_limited",
                    "error": "Xiaohongshu security verification: please wait and retry with browser/CDP.",
                }
            data = extract_from_html(text, strip_tags(text), note.note_id)
            if is_placeholder_copy(data.get("copywriting")) and not data.get("topics"):
                data["copywriting"] = ""
            data["cover"] = self._maybe_download_cover(data.get("cover"), note.row_number)
            return {**data, "status": "ok"}
        except Exception as exc:
            return {"status": "failed", "error": str(exc)}

    def _maybe_download_cover(self, cover_url: Any, row_number: int) -> str:
        cover_url = str(cover_url or "").strip()
        if not cover_url or not self.download_covers or not self.cover_dir:
            return cover_url
        try:
            self.cover_dir.mkdir(parents=True, exist_ok=True)
            headers = {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
                ),
                "Referer": "https://www.xiaohongshu.com/",
            }
            body, content_type = http_get(cover_url, headers=headers)
            content_type = content_type.lower()
            ext = ".jpg"
            if "png" in content_type:
                ext = ".png"
            elif "webp" in content_type:
                ext = ".webp"
            path = self.cover_dir / f"row_{row_number}{ext}"
            path.write_bytes(body)
            return str(path)
        except Exception:
            return cover_url


def http_get(url: str, headers: dict[str, str] | None = None, timeout: int = 20) -> tuple[bytes, str]:
    import urllib.request

    request = urllib.request.Request(url, headers=headers or {})
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read(), response.headers.get("content-type", "")


def safe_page_text(page: Any, timeout: int = 5000) -> str:
    try:
        return page.locator("body").inner_text(timeout=timeout)
    except Exception:
        try:
            return page.evaluate('document.body ? document.body.innerText : ""')
        except Exception:
            return ""


def extract_dom_comments(page: Any, limit: int = 20) -> str:
    try:
        stable_rounds = 0
        last_count = 0
        for _ in range(14):
            expand_more_comments(page)
            scroll_comment_area(page)
            page.wait_for_timeout(900)
            count = page.locator(".comment-item").count()
            if count >= limit:
                break
            if count == last_count:
                stable_rounds += 1
            else:
                stable_rounds = 0
                last_count = count
            if count > 0 and stable_rounds >= 3:
                break
        try:
            page.wait_for_selector(".comment-item", timeout=3000)
        except Exception:
            pass
        items = page.locator(".comment-item")
        count = min(items.count(), limit)
        comments: list[str] = []
        for idx in range(count):
            text = clean_comment_text(items.nth(idx).evaluate("el => el.innerText"))
            if text:
                comments.append(f"{len(comments)+1}. {text}")
        return "\n".join(comments)
    except Exception:
        return ""


def extract_comments_from_payload(payload: Any, limit: int = 20) -> list[str]:
    found: list[str] = []

    def add(item: Any) -> None:
        if len(found) >= limit or not isinstance(item, dict):
            return
        text = first_nonempty(
            item.get("content"),
            item.get("text"),
            item.get("comment"),
            item.get("desc"),
        )
        if not text:
            return
        user = item.get("user_info") or item.get("userInfo") or item.get("user") or {}
        nickname = ""
        if isinstance(user, dict):
            nickname = first_nonempty(user.get("nickname"), user.get("name"), user.get("nickName"))
        cleaned = clean_text(text)
        found.append(f"{nickname + ' / ' if nickname else ''}{cleaned}")

    def walk(obj: Any) -> None:
        if len(found) >= limit:
            return
        if isinstance(obj, dict):
            for key, value in obj.items():
                if key in {"comments", "comment_list", "commentList", "sub_comments", "subComments"} and isinstance(value, list):
                    for item in value:
                        add(item)
                        if isinstance(item, dict):
                            walk(item)
                        if len(found) >= limit:
                            break
                elif isinstance(value, (dict, list)):
                    walk(value)
        elif isinstance(obj, list):
            for item in obj:
                walk(item)
                if len(found) >= limit:
                    break

    walk(payload)
    return found[:limit]


def merge_comments(api_comments: list[str], dom_comments: str, limit: int = 20) -> str:
    merged: list[str] = []
    seen: set[str] = set()
    candidates = list(api_comments)
    candidates.extend(
        re.sub(r"^\d+\.\s*", "", line).strip()
        for line in str(dom_comments or "").splitlines()
        if line.strip()
    )
    for text in candidates:
        normalized = re.sub(r"\s+", " ", text).strip()
        key = comment_dedupe_key(normalized)
        if not normalized or key in seen:
            continue
        seen.add(key)
        merged.append(f"{len(merged)+1}. {normalized}")
        if len(merged) >= limit:
            break
    return "\n".join(merged)


def count_numbered_lines(text: Any) -> int:
    return sum(1 for line in str(text or "").splitlines() if line.strip())


def fetch_signed_note_comments(page: Any, note: SourceNote, limit: int = 20) -> tuple[list[str], str]:
    if not note.note_id:
        return [], "missing note_id"
    if not note.xsec_token:
        return [], "missing xsec_token"
    try:
        sign_with_xhshow = load_media_crawler_signer()
    except Exception as exc:
        return [], f"signer unavailable ({exc})"

    try:
        cookie_str = browser_context_cookie_string(page.context)
    except Exception as exc:
        return [], f"cookie unavailable ({exc})"
    if not cookie_str:
        return [], "empty cookie"

    comments: list[str] = []
    cursor = ""
    has_more = True
    uri = "/api/sns/web/v2/comment/page"
    last_error = ""

    for _ in range(6):
        if len(comments) >= limit or not has_more:
            break
        params = {
            "note_id": note.note_id,
            "cursor": cursor,
            "top_comment_id": "",
            "image_formats": "jpg,webp,avif",
            "xsec_token": note.xsec_token,
        }
        try:
            signed_headers = sign_with_xhshow(
                uri,
                data=params,
                cookie_str=cookie_str,
                method="GET",
            )
            headers = {
                "Accept": "application/json, text/plain, */*",
                "Content-Type": "application/json;charset=UTF-8",
                "Cookie": cookie_str,
                "Origin": "https://www.xiaohongshu.com",
                "Referer": note.url,
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
                ),
                **{key: str(value) for key, value in signed_headers.items() if value},
            }
            query = urllib.parse.urlencode(
                params,
                quote_via=urllib.parse.quote,
                safe=",",
            )
            request = urllib.request.Request(
                f"https://edith.xiaohongshu.com{uri}?{query}",
                headers=headers,
                method="GET",
            )
            with urllib.request.urlopen(request, timeout=20) as response:
                payload = json.loads(response.read().decode("utf-8", errors="replace"))
            if payload.get("success") is False:
                last_error = str(first_nonempty(payload.get("msg"), payload.get("message"), payload.get("code")))
                break
            data = payload.get("data") if isinstance(payload, dict) else {}
            comments.extend(extract_comments_from_payload(payload, limit=limit))
            if isinstance(data, dict):
                cursor = str(data.get("cursor") or "")
                has_more = bool(data.get("has_more") or data.get("hasMore"))
            else:
                has_more = False
            time.sleep(0.4)
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")[:300]
            last_error = f"http {exc.code}: {body}"
            break
        except Exception as exc:
            last_error = str(exc)
            break

    merged = merge_comment_lists(comments, limit=limit)
    return merged, last_error


def fetch_signed_author_profile(page: Any, author_id: str) -> tuple[dict[str, Any], str]:
    author_id = str(author_id or "").strip()
    if not author_id:
        return {}, "missing author_id"
    try:
        sign_with_xhshow = load_media_crawler_signer()
        cookie_str = browser_context_cookie_string(page.context)
    except Exception as exc:
        return {}, str(exc)
    if not cookie_str:
        return {}, "empty cookie"

    uri = "/api/sns/web/v1/user/otherinfo"
    params = {"target_user_id": author_id}
    try:
        signed_headers = sign_with_xhshow(
            uri,
            data=params,
            cookie_str=cookie_str,
            method="GET",
        )
        headers = {
            "Accept": "application/json, text/plain, */*",
            "Cookie": cookie_str,
            "Origin": "https://www.xiaohongshu.com",
            "Referer": f"https://www.xiaohongshu.com/user/profile/{author_id}",
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
            ),
            **{key: str(value) for key, value in signed_headers.items() if value},
        }
        query = urllib.parse.urlencode(params)
        request = urllib.request.Request(
            f"https://edith.xiaohongshu.com{uri}?{query}",
            headers=headers,
            method="GET",
        )
        with urllib.request.urlopen(request, timeout=20) as response:
            payload = json.loads(response.read().decode("utf-8", errors="replace"))
        if payload.get("success") is False:
            return {}, str(first_nonempty(payload.get("msg"), payload.get("message"), payload.get("code")))
        data = payload.get("data") if isinstance(payload, dict) else {}
        basic = data.get("basic_info") if isinstance(data, dict) else {}
        nickname = ""
        if isinstance(basic, dict):
            nickname = str(first_nonempty(basic.get("nickname"), basic.get("name")))
        return {
            "author_nickname": nickname,
            "fans_count": find_fans_count(payload),
        }, ""
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")[:300]
        return {}, f"http {exc.code}: {body}"
    except Exception as exc:
        return {}, str(exc)


def merge_comment_lists(comments: list[str], limit: int = 20) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for text in comments:
        normalized = re.sub(r"\s+", " ", re.sub(r"^\d+\.\s*", "", str(text or ""))).strip()
        key = comment_dedupe_key(normalized)
        if not normalized or key in seen:
            continue
        seen.add(key)
        merged.append(normalized)
        if len(merged) >= limit:
            break
    return merged


def browser_context_cookie_string(context: Any) -> str:
    cookies = context.cookies(["https://www.xiaohongshu.com", "https://edith.xiaohongshu.com"])
    parts = []
    seen: set[str] = set()
    for cookie in cookies:
        name = str(cookie.get("name") or "").strip()
        value = str(cookie.get("value") or "")
        if not name or name in seen:
            continue
        seen.add(name)
        parts.append(f"{name}={value}")
    return "; ".join(parts)


def load_media_crawler_signer() -> Any:
    return sign_with_xhshow_light


def patch_xhshow_get_signature() -> None:
    from xhshow.core.crypto import CryptoProcessor

    if getattr(CryptoProcessor.build_payload_array, "_xhs_note_agent_patched", False):
        return
    original_build = CryptoProcessor.build_payload_array

    def patched_build(
        self: Any,
        hex_parameter: Any,
        a1_value: Any,
        app_identifier: str = "xhs-pc-web",
        string_param: str = "",
        timestamp: Any = None,
        sign_state: Any = None,
    ) -> Any:
        payload = original_build(
            self,
            hex_parameter,
            a1_value,
            app_identifier,
            string_param,
            timestamp,
            sign_state,
        )
        if "{" not in str(string_param):
            correct_md5_hex = hashlib.md5(str(string_param).encode("utf-8")).hexdigest()
            correct_md5_bytes = [int(correct_md5_hex[i : i + 2], 16) for i in range(0, 32, 2)]
            seed_byte = payload[4]
            ts_bytes = payload[8:16]
            correct_a3_hash = self._custom_hash_v2(list(ts_bytes) + correct_md5_bytes)
            for i in range(16):
                payload[128 + i] = correct_a3_hash[i] ^ seed_byte
        return payload

    patched_build._xhs_note_agent_patched = True
    CryptoProcessor.build_payload_array = patched_build


def build_xhs_sign_string(uri: str, data: dict[str, Any] | None = None) -> str:
    if not data:
        return uri
    parts = []
    for key, value in data.items():
        if isinstance(value, list):
            value_str = ",".join(str(v) for v in value)
        elif value is None:
            value_str = ""
        else:
            value_str = str(value)
        parts.append(f"{key}={urllib.parse.quote(value_str, safe=',')}")
    return f"{uri}?{'&'.join(parts)}"


def sign_with_xhshow_light(
    uri: str,
    data: dict[str, Any] | None = None,
    cookie_str: str = "",
    method: str = "GET",
) -> dict[str, str]:
    from xhshow import Xhshow

    if method.upper() != "GET":
        raise ValueError("Only GET signing is implemented for comment pagination")
    patch_xhshow_get_signature()
    xhshow_client = Xhshow()
    content_string = build_xhs_sign_string(uri, data)
    cookie_dict = xhshow_client._parse_cookies(cookie_str)
    a1_value = cookie_dict.get("a1", "")
    ts = time.time()
    d_value = hashlib.md5(content_string.encode("utf-8")).hexdigest()
    payload_array = xhshow_client.crypto_processor.build_payload_array(
        d_value,
        a1_value,
        "xhs-pc-web",
        content_string,
        ts,
    )
    xor_result = xhshow_client.crypto_processor.bit_ops.xor_transform_array(payload_array)
    config = xhshow_client.config
    x3_b64 = xhshow_client.crypto_processor.b64encoder.encode_x3(
        xor_result[: config.PAYLOAD_LENGTH]
    )
    sig_data = config.SIGNATURE_DATA_TEMPLATE.copy()
    sig_data["x3"] = config.X3_PREFIX + x3_b64
    x_s = config.XYS_PREFIX + xhshow_client.crypto_processor.b64encoder.encode(
        json.dumps(sig_data, separators=(",", ":"), ensure_ascii=False)
    )
    return {
        "x-s": x_s,
        "x-s-common": xhshow_client.sign_xs_common(cookie_dict),
        "x-t": str(xhshow_client.get_x_t(ts)),
        "x-b3-traceid": "".join(random.choice("abcdef0123456789") for _ in range(16)),
    }


def comment_dedupe_key(text: str) -> str:
    key = clean_text(text)
    key = re.sub(r"\s*/\s*\d+(?:\s*/\s*\d+)*\s*$", "", key)
    key = re.sub(r"\[[^\]]+R\]", "", key)
    key = re.sub(r"\s+", "", key)
    return key.lower()


def scroll_comment_area(page: Any) -> None:
    page.evaluate(
        """
        () => {
          const explicit = document.querySelector('.note-scroller');
          if (explicit) {
            explicit.scrollTop = explicit.scrollTop + Math.max(700, explicit.clientHeight * 0.9);
            return;
          }
          const candidates = Array.from(document.querySelectorAll('*'))
            .filter(el => {
              const s = window.getComputedStyle(el);
              return /(auto|scroll)/.test(s.overflowY || '') && el.scrollHeight > el.clientHeight + 80;
            })
            .sort((a, b) => b.scrollHeight - a.scrollHeight);
          const comment = candidates.find(el => el.innerText && el.innerText.includes('评论'));
          if (comment) {
            comment.scrollTop = comment.scrollTop + Math.max(600, comment.clientHeight * 0.85);
          } else {
            window.scrollBy(0, 1200);
          }
        }
        """
    )


def expand_more_comments(page: Any) -> None:
    try:
        page.evaluate(
            """
            () => {
              const nodes = Array.from(document.querySelectorAll('div, span, button'));
              const targets = nodes.filter(el => {
                const text = (el.innerText || '').trim();
                return /展开\\s*\\d+\\s*条回复|展开更多|查看更多|更多回复/.test(text);
              }).slice(0, 6);
              for (const el of targets) {
                const clickable = el.closest('button') || el;
                try { clickable.click(); } catch (e) {}
              }
            }
            """
        )
    except Exception:
        pass


def clean_comment_text(text: str) -> str:
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    stop_words = {"赞", "回复", "置顶评论", "作者"}
    filtered = []
    for line in lines:
        if line in stop_words:
            continue
        if re.match(r"\d{2}-\d{2}", line):
            continue
        filtered.append(line)
    return " / ".join(filtered[:4])


def fetch_author_profile(page: Any, author_url: str) -> dict[str, Any]:
    try:
        page.goto(author_url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(2500)
        html_text = page.content()
        body_text = page.evaluate('document.body ? document.body.innerText : ""')
        if is_security_verification(body_text, html_text):
            return {}
        return extract_author_profile(body_text, html_text)
    except Exception:
        return {}


def extract_author_profile(body_text: str, html_text: str) -> dict[str, Any]:
    initial = extract_initial_state(html_text)
    nickname = nested_find(initial, ["nickname", "nickName", "nick_name", "name"])
    fans = find_fans_count(initial)
    if fans is None:
        fans = parse_fans_from_text(body_text)
    if not nickname:
        lines = [line.strip() for line in str(body_text or "").splitlines() if line.strip()]
        if lines:
            nickname = lines[0]
    return {"author_nickname": nickname, "fans_count": fans}


def find_fans_count(obj: Any) -> int | None:
    fan_keys = {
        "fans",
        "fansCount",
        "fans_count",
        "followerCount",
        "follower_count",
        "followers",
        "followersCount",
        "followers_count",
    }
    if isinstance(obj, dict):
        interactions = obj.get("interactions")
        if isinstance(interactions, list):
            for item in interactions:
                if not isinstance(item, dict):
                    continue
                marker = f"{item.get('type', '')} {item.get('name', '')}".lower()
                if "fans" in marker or "粉丝" in marker:
                    parsed = parse_int(first_nonempty(item.get("count"), item.get("i18n_count")))
                    if parsed is not None:
                        return parsed
        for key, value in obj.items():
            if key in fan_keys:
                parsed = parse_int(value)
                if parsed is not None:
                    return parsed
            if isinstance(value, (dict, list)):
                found = find_fans_count(value)
                if found is not None:
                    return found
    elif isinstance(obj, list):
        for item in obj:
            found = find_fans_count(item)
            if found is not None:
                return found
    return None


def parse_fans_from_text(text: str) -> int | None:
    compact = re.sub(r"\s+", " ", text or "")
    patterns = [
        r"粉丝\s*([0-9,.]+(?:\.\d+)?万?)",
        r"([0-9,.]+(?:\.\d+)?万?)\s*粉丝",
        r"followers?\s*([0-9,.]+(?:\.\d+)?[kK万]?)",
    ]
    for pattern in patterns:
        match = re.search(pattern, compact, flags=re.I)
        if match:
            return parse_int(match.group(1))
    return None


def find_chrome_executable() -> Path | None:
    candidates = [
        Path("C:/Program Files/Google/Chrome/Application/chrome.exe"),
        Path("C:/Program Files (x86)/Google/Chrome/Application/chrome.exe"),
        Path.home() / "AppData/Local/Google/Chrome/Application/chrome.exe",
        Path("C:/Program Files/Microsoft/Edge/Application/msedge.exe"),
        Path("C:/Program Files (x86)/Microsoft/Edge/Application/msedge.exe"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def strip_tags(text: str) -> str:
    text = re.sub(r"<script.*?</script>", " ", text, flags=re.S | re.I)
    text = re.sub(r"<style.*?</style>", " ", text, flags=re.S | re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    return html.unescape(re.sub(r"\s+", " ", text)).strip()


def is_security_verification(page_text: str, html_text: str) -> bool:
    combined = f"{page_text}\n{html_text}"
    return "\u5b89\u5168\u9a8c\u8bc1" in combined or "\u8bf7\u52ff\u9891\u7e41\u64cd\u4f5c" in combined


def is_missing_note_page(page_text: str, html_text: str) -> bool:
    combined = clean_text(f"{page_text}\n{html_text}")
    return (
        "你访问的页面不见了" in combined
        or "当前内容无法查看" in combined
        or "note is not found" in combined.lower()
    )


def is_placeholder_copy(text: Any) -> bool:
    normalized = clean_text(str(text or ""))
    return normalized in {"\u8fd8\u6ca1\u6709\u7b80\u4ecb", "\u6682\u65e0\u7b80\u4ecb"}


def extract_from_html(html_text: str, page_text: str = "", note_id: str = "") -> dict[str, Any]:
    meta = extract_meta(html_text)
    initial = extract_initial_state(html_text)
    note_card = extract_note_card(initial, note_id)
    text_pool = "\n".join(filter(None, [page_text, meta.get("description", "")]))
    title = first_nonempty(
        note_card.get("title"),
        meta.get("og:title"),
        meta.get("title"),
    )
    copywriting = first_nonempty(
        note_card.get("desc"),
        note_card.get("content"),
        meta.get("description"),
        guess_copywriting(text_pool),
    )
    cover = first_nonempty(
        note_card.get("cover"),
        meta.get("og:image"),
    )
    topics = note_card.get("topics") or "、".join(extract_topics(copywriting or text_pool))
    metrics = note_card.get("metrics") or {}
    return {
        "title": clean_title(title),
        "copywriting": clean_text(copywriting),
        "cover": cover,
        "topics": topics,
        "author_nickname": note_card.get("author_nickname", ""),
        "author_id": note_card.get("author_id", ""),
        "author_url": note_card.get("author_url", ""),
        "fans_count": note_card.get("fans_count"),
        "top_comments": note_card.get("top_comments", ""),
        **metrics,
    }


def extract_meta(html_text: str) -> dict[str, str]:
    meta: dict[str, str] = {}
    title_match = re.search(r"<title[^>]*>(.*?)</title>", html_text, flags=re.S | re.I)
    if title_match:
        meta["title"] = html.unescape(strip_tags(title_match.group(1)))
    for match in re.finditer(r"<meta\s+([^>]+)>", html_text, flags=re.I):
        attrs = {
            k.lower(): html.unescape(v)
            for k, _, v in re.findall(
                r'([\w:-]+)\s*=\s*([\"\'])(.*?)\2', match.group(1)
            )
        }
        key = attrs.get("property") or attrs.get("name")
        content = attrs.get("content")
        if key and content:
            meta[key.lower()] = content
    return meta


def extract_initial_state(html_text: str) -> dict[str, Any]:
    candidates = re.findall(r"window\.__INITIAL_STATE__\s*=\s*({.*?})</script>", html_text, flags=re.S)
    if not candidates:
        candidates = re.findall(r"window\.__INITIAL_STATE__=({.*})</script>", html_text, flags=re.S)
    for raw in candidates:
        raw = raw.replace(":undefined", ":null").replace("undefined", "null")
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            continue
    return {}


def extract_note_card(initial: dict[str, Any], note_id: str) -> dict[str, Any]:
    note = find_note_by_id(initial, note_id) if initial else {}
    if not note:
        return {}
    interact = note.get("interactInfo") or note.get("interact_info") or {}
    metrics = {
        "likes": parse_int(first_nonempty(interact.get("likedCount"), interact.get("liked_count"), interact.get("likeCount"))),
        "collects": parse_int(first_nonempty(interact.get("collectedCount"), interact.get("collected_count"), interact.get("collectCount"))),
        "comments": parse_int(first_nonempty(interact.get("commentCount"), interact.get("comment_count"))),
        "shares": parse_int(first_nonempty(interact.get("shareCount"), interact.get("share_count"))),
    }
    return {
        "title": first_nonempty(note.get("title"), note.get("displayTitle")),
        "desc": first_nonempty(note.get("desc"), note.get("content")),
        "topics": extract_structured_topics(note),
        "cover": extract_cover_url(note),
        **extract_author_info(note),
        "top_comments": extract_top_comments(note, limit=20),
        "metrics": {key: value for key, value in metrics.items() if value is not None},
    }


def extract_author_info(note: dict[str, Any]) -> dict[str, Any]:
    user = (
        note.get("user")
        or note.get("userInfo")
        or note.get("user_info")
        or note.get("author")
        or {}
    )
    if not isinstance(user, dict):
        user = {}
    user_id = first_nonempty(
        user.get("userId"),
        user.get("user_id"),
        user.get("id"),
        note.get("userId"),
        note.get("user_id"),
    )
    nickname = first_nonempty(
        user.get("nickname"),
        user.get("nickName"),
        user.get("nick_name"),
        user.get("name"),
    )
    fans = parse_int(
        first_nonempty(
            user.get("fans"),
            user.get("fansCount"),
            user.get("fans_count"),
            user.get("followerCount"),
            user.get("follower_count"),
        )
    )
    return {
        "author_nickname": nickname,
        "author_id": user_id,
        "author_url": f"https://www.xiaohongshu.com/user/profile/{user_id}" if user_id else "",
        "fans_count": fans,
    }


def extract_top_comments(note: dict[str, Any], limit: int = 20) -> str:
    comments: list[str] = []

    def add_comment(item: Any) -> None:
        if len(comments) >= limit:
            return
        if isinstance(item, dict):
            text = first_nonempty(
                item.get("content"),
                item.get("text"),
                item.get("comment"),
                item.get("desc"),
            )
            user = item.get("user") or item.get("userInfo") or item.get("user_info") or {}
            nickname = ""
            if isinstance(user, dict):
                nickname = first_nonempty(user.get("nickname"), user.get("name"))
            if text:
                comments.append(f"{len(comments)+1}. {nickname + ': ' if nickname else ''}{clean_text(text)}")
        elif isinstance(item, str) and item.strip():
            comments.append(f"{len(comments)+1}. {clean_text(item)}")

    def walk(obj: Any) -> None:
        if len(comments) >= limit:
            return
        if isinstance(obj, dict):
            for key, value in obj.items():
                if key in {"comments", "commentList", "comment_list", "topComments", "top_comments"} and isinstance(value, list):
                    for item in value:
                        add_comment(item)
                        if len(comments) >= limit:
                            break
                elif isinstance(value, (dict, list)):
                    walk(value)
        elif isinstance(obj, list):
            for item in obj:
                walk(item)
                if len(comments) >= limit:
                    break

    walk(note)
    return "\n".join(comments[:limit])


def find_note_by_id(obj: Any, note_id: str) -> dict[str, Any]:
    if not isinstance(obj, (dict, list)):
        return {}
    if isinstance(obj, dict):
        note_map = obj.get("noteDetailMap") or obj.get("note_detail_map")
        if isinstance(note_map, dict) and note_id:
            wrapped = note_map.get(note_id)
            if isinstance(wrapped, dict):
                return wrapped.get("note") or wrapped.get("noteCard") or wrapped.get("note_card") or wrapped
        if note_id and str(obj.get("noteId") or obj.get("note_id") or obj.get("id") or "") == note_id:
            return obj
        for value in obj.values():
            found = find_note_by_id(value, note_id)
            if found:
                return found
    else:
        for value in obj:
            found = find_note_by_id(value, note_id)
            if found:
                return found
    return {}


def extract_structured_topics(note: dict[str, Any]) -> str:
    topic_names: list[str] = []
    for candidate in (
        note.get("tagList"),
        note.get("tag_list"),
        note.get("hashTag"),
        note.get("hash_tag"),
        note.get("topics"),
    ):
        if not isinstance(candidate, list):
            continue
        for item in candidate:
            if isinstance(item, dict):
                name = first_nonempty(item.get("name"), item.get("tagName"), item.get("tag_name"))
            else:
                name = str(item)
            if name and name not in topic_names:
                topic_names.append(name)
    if not topic_names:
        topic_names = extract_topics(first_nonempty(note.get("desc"), note.get("content")))
    return "、".join(topic_names)


def extract_cover_url(note: dict[str, Any]) -> str:
    image_list = note.get("imageList") or note.get("image_list") or []
    if isinstance(image_list, list) and image_list:
        first = image_list[0]
        if isinstance(first, dict):
            for key in ("urlDefault", "url_default", "urlPre", "url_pre", "url"):
                if first.get(key):
                    return str(first[key])
            info_list = first.get("infoList") or first.get("info_list") or []
            if isinstance(info_list, list):
                for info in info_list:
                    if isinstance(info, dict) and info.get("url"):
                        return str(info["url"])
    cover = note.get("cover")
    if isinstance(cover, dict):
        return first_nonempty(cover.get("url"), cover.get("urlDefault"), cover.get("url_default"))
    return str(cover or "")


def nested_find(value: Any, keys: list[str]) -> str:
    found: list[str] = []

    def walk(obj: Any) -> None:
        if isinstance(obj, dict):
            for key, item in obj.items():
                if key in keys and isinstance(item, (str, int, float)):
                    found.append(str(item))
                else:
                    walk(item)
        elif isinstance(obj, list):
            for item in obj:
                walk(item)

    walk(value)
    return first_nonempty(*found)


def first_nonempty(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def clean_title(title: str) -> str:
    title = clean_text(title)
    title = re.sub(r"\s*-\s*小红书.*$", "", title)
    return title


def clean_text(text: str) -> str:
    text = html.unescape(str(text or ""))
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_excel_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    return re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", "", value)


def guess_copywriting(text: str) -> str:
    lines = [line.strip() for line in text.splitlines() if len(line.strip()) > 15]
    return max(lines, key=len) if lines else clean_text(text[:1000])


def extract_topics(text: str) -> list[str]:
    topics = re.findall(r"#([^#\s，,。；;！!？?\[\]【】]+)", text or "")
    seen: set[str] = set()
    result = []
    for topic in topics:
        if topic not in seen:
            seen.add(topic)
            result.append(topic)
    return result


def extract_metrics(text: str) -> dict[str, int | None]:
    patterns = {
        "likes": r"(?:点赞|赞)\s*([0-9.]+万?|[0-9.]+k?)",
        "collects": r"(?:收藏)\s*([0-9.]+万?|[0-9.]+k?)",
        "comments": r"(?:评论)\s*([0-9.]+万?|[0-9.]+k?)",
        "shares": r"(?:分享|转发)\s*([0-9.]+万?|[0-9.]+k?)",
    }
    return {
        field: parse_int(match.group(1)) if (match := re.search(pattern, text, re.I)) else None
        for field, pattern in patterns.items()
    }


def analyze(note: SourceNote, crawled: dict[str, Any]) -> NoteResult:
    crawled_title = normalize_crawled_title(crawled.get("title"))
    title = first_nonempty(crawled_title, note.title)
    copywriting = str(crawled.get("copywriting") or "")
    topics = str(crawled.get("topics") or "")
    author_nickname = str(crawled.get("author_nickname") or "")
    author_id = first_nonempty(crawled.get("author_id"), note.author_id)
    author_url = first_nonempty(crawled.get("author_url"), note.author_url)
    fans_count = parse_int(crawled.get("fans_count"))
    top_comments = str(crawled.get("top_comments") or "")
    likes = note.likes if note.likes is not None else parse_int(crawled.get("likes"))
    collects = note.collects if note.collects is not None else parse_int(crawled.get("collects"))
    comments = note.comments if note.comments is not None else parse_int(crawled.get("comments"))
    shares = note.shares if note.shares is not None else parse_int(crawled.get("shares"))
    total_interactions = sum(v or 0 for v in [likes, collects, comments, shares])
    pgy_image_price = note.pgy_image_price or note.pgy_price
    pgy_video_price = note.pgy_video_price
    pgy_price = pgy_image_price
    cpe = round(pgy_price / total_interactions, 4) if pgy_price and total_interactions else None
    image_cpe = round(pgy_image_price / total_interactions, 4) if pgy_image_price and total_interactions else None
    video_cpe = round(pgy_video_price / total_interactions, 4) if pgy_video_price and total_interactions else None

    content = f"{title} {copywriting} {topics}"
    content_type = classify_content(content)
    content_group = classify_content_group(content_type, content)
    selling_points = infer_selling_points(content)
    title_pattern = classify_title(title)
    engagement_type = classify_engagement(likes, collects, comments, shares)
    creative_advice = build_creative_advice(
        content_type, title_pattern, selling_points, engagement_type, content
    )
    audience_strategy = build_audience_strategy(content, content_type, engagement_type)

    return NoteResult(
        row_number=note.row_number,
        title=title,
        url=note.url,
        cover=str(crawled.get("cover") or ""),
        copywriting=copywriting,
        topics=topics,
        author_nickname=author_nickname,
        author_id=author_id,
        author_url=author_url,
        fans_count=fans_count,
        top_comments=top_comments,
        likes=likes,
        collects=collects,
        comments=comments,
        shares=shares,
        total_interactions=total_interactions,
        pgy_url=note.pgy_url,
        pgy_price=pgy_price,
        pgy_image_price=pgy_image_price,
        pgy_video_price=pgy_video_price,
        cpe=cpe,
        image_cpe=image_cpe,
        video_cpe=video_cpe,
        content_type=content_type,
        content_group=content_group,
        title_pattern=title_pattern,
        selling_points=selling_points,
        engagement_type=engagement_type,
        creative_advice=creative_advice,
        audience_strategy=audience_strategy,
        status=str(crawled.get("status") or "offline"),
        error=str(crawled.get("error") or ""),
    )


class LlmAnalyzer:
    def __init__(
        self,
        api_key: str,
        base_url: str,
        model: str,
        timeout: int = 60,
        temperature: float = 0.3,
    ):
        self.api_key = api_key
        self.base_url = base_url.rstrip("/")
        self.model = model
        self.timeout = timeout
        self.temperature = temperature

    @classmethod
    def from_args(cls, args: argparse.Namespace) -> "LlmAnalyzer":
        api_key = args.llm_api_key or os.getenv("LLM_API_KEY") or os.getenv("OPENAI_API_KEY")
        base_url = (
            args.llm_base_url
            or os.getenv("LLM_BASE_URL")
            or os.getenv("OPENAI_BASE_URL")
            or "https://api.openai.com/v1"
        )
        model = args.llm_model or os.getenv("LLM_MODEL") or "gpt-4.1-mini"
        if not api_key:
            raise ValueError("Missing LLM API key. Set LLM_API_KEY or pass --llm-api-key.")
        return cls(
            api_key=api_key,
            base_url=base_url,
            model=model,
            timeout=args.llm_timeout,
            temperature=args.llm_temperature,
        )

    def improve(self, result: NoteResult) -> NoteResult:
        payload = {
            "model": self.model,
            "temperature": self.temperature,
            "response_format": {"type": "json_object"},
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "你是小红书内容策略分析师。请基于笔记标题、正文、话题和互动数据，"
                        "输出严格 JSON，不要 Markdown。字段包括：content_type, title_pattern, "
                        "selling_points, engagement_type, creative_advice, audience_strategy。"
                        "建议要具体、可执行，适合家电/空调内容种草。"
                    ),
                },
                {
                    "role": "user",
                    "content": json.dumps(
                        {
                            "标题": result.title,
                            "文案": result.copywriting,
                            "话题": result.topics,
                            "达人昵称": result.author_nickname,
                            "达人ID": result.author_id,
                            "粉丝量": result.fans_count,
                            "评论区前20条": result.top_comments,
                            "点赞数": result.likes,
                            "收藏数": result.collects,
                            "评论数": result.comments,
                            "分享数": result.shares,
                            "总互动量": result.total_interactions,
                            "蒲公英报价": result.pgy_price,
                            "蒲公英图文报价": result.pgy_image_price,
                            "蒲公英视频报价": result.pgy_video_price,
                            "CPE": result.cpe,
                            "图文CPE": result.image_cpe,
                            "视频CPE": result.video_cpe,
                            "规则初判": {
                                "内容类型": result.content_type,
                                "内容类型分组": result.content_group,
                                "标题结构": result.title_pattern,
                                "核心卖点": result.selling_points,
                                "互动倾向": result.engagement_type,
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ],
        }
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        request = urllib.request.Request(
            f"{self.base_url}/chat/completions",
            data=data,
            headers={
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=self.timeout) as response:
                raw = response.read().decode("utf-8", errors="replace")
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"HTTP {exc.code}: {body}") from exc
        completion = json.loads(raw)
        content = completion["choices"][0]["message"]["content"]
        parsed = parse_llm_json(content)
        result.content_type = str(parsed.get("content_type") or result.content_type)
        result.title_pattern = str(parsed.get("title_pattern") or result.title_pattern)
        result.selling_points = str(parsed.get("selling_points") or result.selling_points)
        result.engagement_type = str(parsed.get("engagement_type") or result.engagement_type)
        result.creative_advice = str(parsed.get("creative_advice") or result.creative_advice)
        result.audience_strategy = str(parsed.get("audience_strategy") or result.audience_strategy)
        result.llm_status = "ok"
        result.llm_model = self.model
        return result


def parse_llm_json(content: str) -> dict[str, Any]:
    content = content.strip()
    if content.startswith("```"):
        content = re.sub(r"^```(?:json)?\s*", "", content)
        content = re.sub(r"\s*```$", "", content)
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", content, flags=re.S)
        if match:
            return json.loads(match.group(0))
        raise


def apply_llm_analysis(results: list[NoteResult], analyzer: LlmAnalyzer) -> list[NoteResult]:
    for idx, result in enumerate(results, start=1):
        if result.status in {"rate_limited", "skipped_rate_limit", "failed"}:
            result.llm_status = f"skipped:{result.status}"
            result.llm_model = analyzer.model
            continue
        safe_print(f"[LLM {idx}/{len(results)}] analyze row {result.row_number}")
        try:
            analyzer.improve(result)
        except Exception as exc:
            result.llm_status = f"failed:{exc}"
            result.llm_model = analyzer.model
    return results


class PgyCrawler:
    def __init__(
        self,
        cdp_url: str,
        timeout_ms: int = 30000,
        delay: float = 3.0,
        safe_mode: bool = False,
        max_retries: int = 1,
    ):
        self.cdp_url = cdp_url
        self.timeout_ms = timeout_ms
        self.delay = delay
        self.safe_mode = safe_mode
        self.max_retries = max(1, max_retries)

    def enrich(self, results: list[NoteResult]) -> list[NoteResult]:
        try:
            from playwright.sync_api import sync_playwright

            with sync_playwright() as p:
                browser = p.chromium.connect_over_cdp(self.cdp_url)
                context = browser.contexts[0] if browser.contexts else browser.new_context()
                page = context.new_page()
                for idx, result in enumerate(results, start=1):
                    if result.status not in {"ok", "offline"}:
                        continue
                    target = self._target_url(result)
                    if not target:
                        continue
                    safe_print(f"[PGY {idx}/{len(results)}] {target}")
                    last_exc = None
                    for attempt in range(1, self.max_retries + 1):
                        try:
                            page.goto(target, wait_until="domcontentloaded", timeout=self.timeout_ms)
                            page.wait_for_timeout(2500 if not self.safe_mode else 5000)
                            text = page.locator("body").inner_text(timeout=5000)
                            parsed = parse_pgy_page_text(text)
                            if parsed.get("image_price") is not None:
                                result.pgy_image_price = parsed["image_price"]
                            if parsed.get("video_price") is not None:
                                result.pgy_video_price = parsed["video_price"]
                            if parsed.get("price") is not None and result.pgy_image_price is None:
                                result.pgy_image_price = parsed["price"]
                            if result.pgy_image_price is not None:
                                result.pgy_price = result.pgy_image_price
                            if parsed.get("fans_count") is not None and result.fans_count is None:
                                result.fans_count = parsed["fans_count"]
                            if not result.pgy_url:
                                result.pgy_url = target
                            result.cpe = calc_cpe(result.pgy_price, result.total_interactions)
                            result.image_cpe = calc_cpe(result.pgy_image_price, result.total_interactions)
                            result.video_cpe = calc_cpe(result.pgy_video_price, result.total_interactions)
                            last_exc = None
                            break
                        except Exception as exc:
                            last_exc = exc
                            if attempt < self.max_retries:
                                backoff = self.delay * (1.5 if self.safe_mode else 1.0)
                                time.sleep(backoff)
                    if last_exc is not None:
                        result.error = append_error(result.error, f"PGY failed: {last_exc}")
                    self._sleep_between_requests()
                page.close()
                return results
        except Exception as exc:
            for result in results:
                result.error = append_error(result.error, f"PGY unavailable: {exc}")
            return results

    def _target_url(self, result: NoteResult) -> str:
        if result.pgy_url:
            return result.pgy_url
        if result.author_id:
            return (
                "https://pgy.xiaohongshu.com/solar/pre-trade/blogger-detail/"
                f"{urllib.parse.quote(result.author_id)}?fromRoute=Advertiser_Kol&source=Advertiser_Kol"
            )
        if result.author_url:
            return f"https://pgy.xiaohongshu.com/solar/search/kol?keyword={urllib.parse.quote(result.author_url)}"
        return ""

    def _sleep_between_requests(self) -> None:
        extra = random.uniform(1.5, 4.0) if self.safe_mode else random.uniform(0.0, 0.8)
        time.sleep(self.delay + extra)


def parse_pgy_page_text(text: str) -> dict[str, Any]:
    compact = re.sub(r"\s+", " ", text or "")
    price = None
    image_price = None
    video_price = None
    image_match = re.search(r"图文笔记一口价\s*[￥¥]?\s*([0-9,.]+(?:\.\d+)?)", compact)
    if image_match:
        image_price = parse_float(image_match.group(1))
    video_match = re.search(r"视频笔记一口价\s*[￥¥]?\s*([0-9,.]+(?:\.\d+)?)", compact)
    if video_match:
        video_price = parse_float(video_match.group(1))
    price_patterns = [
        r"(?:图文|笔记|报价|合作报价|价格)[^\d￥¥]{0,12}[￥¥]?\s*([0-9,.]+(?:\.\d+)?(?:万|w)?)",
        r"[￥¥]\s*([0-9,.]+(?:\.\d+)?(?:万|w)?)",
    ]
    for pattern in price_patterns:
        match = re.search(pattern, compact)
        if match:
            price = parse_float(match.group(1))
            break
    if image_price is None:
        image_price = price
    fans = None
    fans_match = re.search(r"(?:粉丝|粉丝量)[^\d]{0,8}([0-9,.]+(?:\.\d+)?(?:万|w)?)", compact)
    if fans_match:
        fans = parse_int(fans_match.group(1))
    return {
        "price": price,
        "image_price": image_price,
        "video_price": video_price,
        "fans_count": fans,
    }


def calc_cpe(price: float | None, total_interactions: int | None) -> float | None:
    if not price or not total_interactions:
        return None
    return round(float(price) / float(total_interactions), 4)


def append_error(existing: str, message: str) -> str:
    return f"{existing} | {message}" if existing else message


def normalize_crawled_title(title: Any) -> str:
    title_text = clean_title(str(title or ""))
    noise_markers = ["小红书_沪ICP备", "小红书 - 你的生活指南"]
    if any(marker in title_text for marker in noise_markers):
        return ""
    return title_text


def classify_content(content: str) -> str:
    rules = [
        ("选购指南", ["指南", "怎么选", "选购", "避坑", "测评", "横评"]),
        ("家庭场景种草", ["孩子", "宝宝", "老人", "卧室", "客厅", "全家", "母婴"]),
        ("痛点解决", ["后悔", "终于", "救了", "不再", "解决", "难受", "费电"]),
        ("技术卖点", ["ai", "省电", "新风", "静音", "除菌", "风感", "能效", "智能"]),
        ("体验分享", ["用了", "入手", "真实", "体验", "分享", "反馈"]),
    ]
    lower = content.lower()
    hits = [name for name, words in rules if any(word in lower for word in words)]
    return " / ".join(hits[:2]) if hits else "泛种草内容"


def classify_content_group(content_type: str, content: str) -> str:
    if "选购指南" in content_type:
        return "攻略/选购指南"
    if "家庭场景" in content_type:
        return "家庭场景种草"
    if "痛点解决" in content_type:
        return "痛点解决/体验改善"
    if "技术卖点" in content_type:
        return "技术卖点解释"
    if "体验分享" in content_type:
        return "真实体验分享"
    if any(word in content.lower() for word in ["awe", "展会", "发布会"]):
        return "节点/事件内容"
    return "泛种草内容"


def infer_selling_points(content: str) -> str:
    points = [
        ("省电节能", ["省电", "电费", "能效", "节能"]),
        ("舒适风感", ["风感", "不直吹", "柔风", "舒适"]),
        ("母婴/儿童友好", ["孩子", "宝宝", "母婴", "儿童"]),
        ("空气健康", ["新风", "除菌", "净化", "空气", "甲醛"]),
        ("智能体验", ["ai", "智能", "自动", "语音"]),
        ("安静睡眠", ["静音", "睡眠", "卧室", "安静"]),
        ("高性价比", ["性价比", "价格", "划算", "预算"]),
    ]
    lower = content.lower()
    matched = [name for name, words in points if any(word in lower for word in words)]
    return "、".join(matched[:4]) if matched else "需从正文中进一步提炼产品利益点"


def classify_title(title: str) -> str:
    patterns = []
    if re.search(r"\d", title):
        patterns.append("数字/年份增强可信度")
    if any(word in title for word in ["后悔", "早买", "踩坑", "避坑"]):
        patterns.append("反差痛点型")
    if any(word in title for word in ["指南", "怎么选", "选购"]):
        patterns.append("攻略指南型")
    if any(word in title for word in ["孩子", "宝宝", "老人", "卧室"]):
        patterns.append("场景人群型")
    if any(mark in title for mark in ["！", "!", "？", "?"]):
        patterns.append("强情绪钩子")
    return "、".join(patterns) if patterns else "直接种草型"


def classify_engagement(
    likes: int | None, collects: int | None, comments: int | None, shares: int | None
) -> str:
    likes = likes or 0
    collects = collects or 0
    comments = comments or 0
    shares = shares or 0
    total = likes + collects + comments + shares
    if total <= 0:
        return "数据不足"
    collect_ratio = collects / total
    comment_ratio = comments / total
    share_ratio = shares / total
    if collect_ratio >= 0.35:
        return "高收藏实用型"
    if comment_ratio >= 0.12:
        return "高讨论决策型"
    if share_ratio >= 0.08:
        return "高分享传播型"
    return "点赞共鸣型"


def build_creative_advice(
    content_type: str,
    title_pattern: str,
    selling_points: str,
    engagement_type: str,
    content: str,
) -> str:
    advice = [
        f"延续「{title_pattern}」标题结构，首屏直接给出场景痛点和结果承诺。",
        f"封面建议突出「{selling_points}」中的第一优先卖点，采用真实家居/使用场景 + 3到7字利益点大字。",
    ]
    if "高收藏" in engagement_type or "选购指南" in content_type:
        advice.append("增加参数对比、清单式结论、避坑表述，做成可收藏的选购工具型内容。")
    if "家庭" in content_type or any(word in content for word in ["孩子", "宝宝"]):
        advice.append("强化家庭成员视角，例如孩子睡眠、老人舒适、卧室夜间使用，降低硬广感。")
    if "技术卖点" in content_type:
        advice.append("技术点不要单独讲参数，建议用「问题-体验变化-证据」结构转译成用户利益。")
    return " ".join(advice)


def build_audience_strategy(content: str, content_type: str, engagement_type: str) -> str:
    groups = []
    if any(word in content for word in ["孩子", "宝宝", "母婴", "儿童"]):
        groups.append("有孩家庭/母婴人群：关注健康风感、睡眠舒适、空气安全。")
    if any(word in content for word in ["省电", "能效", "电费"]):
        groups.append("精打细算型家电换新人群：关注长期电费、一级能效、真实使用成本。")
    if any(word in content for word in ["选购", "指南", "避坑", "测评"]):
        groups.append("装修换新/家电决策人群：关注型号对比、避坑参数、一步到位购买建议。")
    if any(word in content for word in ["ai", "智能", "新风", "除菌"]):
        groups.append("品质升级人群：关注智能体验、空气健康、舒适度升级。")
    if not groups:
        groups.append("泛家电兴趣人群：先用场景痛点内容筛选高意向用户，再用参数/测评内容承接转化。")
    groups.append(f"互动承接：当前偏「{engagement_type}」，投放/运营上优先匹配相同意图的人群包和搜索词。")
    return " ".join(groups)


def write_outputs(results: list[NoteResult], output_path: Path, embed_covers: bool = False) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path = output_path.with_suffix(".csv")
    columns = [
        ("row_number", "源表行号"),
        ("title", "标题"),
        ("url", "笔记链接"),
        ("cover", "封面"),
        ("copywriting", "文案"),
        ("topics", "话题"),
        ("author_nickname", "达人昵称"),
        ("author_id", "达人ID"),
        ("author_url", "达人链接"),
        ("fans_count", "粉丝量"),
        ("top_comments", "评论区前20条"),
        ("likes", "点赞数"),
        ("collects", "收藏数"),
        ("comments", "评论数"),
        ("shares", "分享数"),
        ("total_interactions", "总互动量"),
        ("pgy_url", "蒲公英链接"),
        ("pgy_price", "蒲公英报价"),
        ("pgy_image_price", "蒲公英图文报价"),
        ("pgy_video_price", "蒲公英视频报价"),
        ("cpe", "CPE"),
        ("image_cpe", "图文CPE"),
        ("video_cpe", "视频CPE"),
        ("content_type", "内容类型"),
        ("content_group", "内容类型分组"),
        ("title_pattern", "标题结构"),
        ("selling_points", "核心卖点"),
        ("engagement_type", "互动倾向"),
        ("creative_advice", "创意建议"),
        ("audience_strategy", "人群圈选策略"),
        ("llm_status", "LLM状态"),
        ("llm_model", "LLM模型"),
        ("status", "采集状态"),
        ("error", "异常信息"),
    ]
    fieldnames = [label for _, label in columns]
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for result in results:
            data = asdict(result)
            writer.writerow({label: clean_excel_value(data.get(key)) for key, label in columns})
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "笔记分析结果"
        headers = fieldnames
        ws.append(headers)
        for result in results:
            data = asdict(result)
            ws.append([clean_excel_value(data.get(key)) for key, _ in columns])
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        widths = {
            "标题": 34,
            "笔记链接": 46,
            "封面": 24 if embed_covers else 40,
            "文案": 60,
            "话题": 30,
            "达人昵称": 18,
            "达人ID": 22,
            "达人链接": 42,
            "评论区前20条": 60,
            "蒲公英链接": 42,
            "蒲公英图文报价": 16,
            "蒲公英视频报价": 16,
            "图文CPE": 14,
            "视频CPE": 14,
            "创意建议": 60,
            "人群圈选策略": 60,
            "LLM状态": 26,
            "LLM模型": 22,
            "异常信息": 36,
        }
        for col_idx, header in enumerate(headers, start=1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = widths.get(header, 16)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if embed_covers:
            embed_cover_images(ws, results, cover_col=4, start_row=2)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(output_path)
    except Exception as exc:
        safe_print(f"Failed to write xlsx, csv is available: {exc}")


def embed_cover_images(ws: Any, results: list[NoteResult], cover_col: int, start_row: int) -> None:
    from openpyxl.drawing.image import Image as ExcelImage

    # 120px roughly maps to 90pt row height in Excel.
    thumb_px = 120
    converted_dir = Path("outputs/embedded_covers")
    converted_dir.mkdir(parents=True, exist_ok=True)
    for offset, result in enumerate(results):
        row_idx = start_row + offset
        cover_path = Path(str(result.cover or ""))
        if not cover_path.exists() or not cover_path.is_file():
            continue
        try:
            image_path = prepare_excel_image_path(cover_path, converted_dir)
            image = ExcelImage(str(image_path))
            width = image.width or thumb_px
            height = image.height or thumb_px
            scale = min(thumb_px / width, thumb_px / height)
            image.width = int(width * scale)
            image.height = int(height * scale)
            ws.add_image(image, ws.cell(row=row_idx, column=cover_col).coordinate)
            ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 15, 95)
        except Exception as exc:
            ws.cell(row=row_idx, column=cover_col).value = f"{result.cover} (embed failed: {exc})"


def prepare_excel_image_path(path: Path, converted_dir: Path) -> Path:
    if path.suffix.lower() in {".png", ".jpg", ".jpeg", ".bmp", ".gif"}:
        return path
    from PIL import Image

    target = converted_dir / f"{path.stem}.png"
    if target.exists() and target.stat().st_mtime >= path.stat().st_mtime:
        return target
    with Image.open(path) as img:
        if img.mode not in {"RGB", "RGBA"}:
            img = img.convert("RGBA")
        img.save(target, "PNG")
    return target


def run(args: argparse.Namespace) -> None:
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    notes = load_source_notes(input_path)
    if args.limit:
        notes = notes[: args.limit]
    safe_print(f"Loaded {len(notes)} notes from 笔记明细")
    if not notes:
        raise SystemExit("No note URLs found. Expected column: 笔记url / 笔记链接 / 链接")

    if args.no_crawl:
        crawled = [{"status": "offline"} for _ in notes]
    else:
        crawler = XhsCrawler(
            headless=args.headless,
            profile=Path(args.profile).resolve() if args.profile else None,
            download_covers=args.download_covers,
            cover_dir=output_path.parent / "covers",
            browser_executable=Path(args.browser_executable).resolve()
            if args.browser_executable
            else None,
            cdp_url=args.cdp_url,
            login_first=args.login_first,
            crawl_delay=args.crawl_delay,
            stop_on_rate_limit=not args.no_stop_on_rate_limit,
            rate_limit_cooldown=args.rate_limit_cooldown,
            comment_api=not args.no_comment_api,
        )
        crawled = crawler.crawl_many(notes)
    results = [analyze(note, data) for note, data in zip(notes, crawled)]
    if args.crawl_pgy:
        if not args.cdp_url:
            raise SystemExit("--crawl-pgy requires --cdp-url so it can reuse your logged-in browser.")
        results = PgyCrawler(
            cdp_url=args.cdp_url,
            timeout_ms=args.pgy_timeout,
            delay=args.pgy_delay,
            safe_mode=args.pgy_safe_mode,
            max_retries=args.pgy_max_retries,
        ).enrich(results)
    if args.use_llm:
        analyzer = LlmAnalyzer.from_args(args)
        results = apply_llm_analysis(results, analyzer)
    write_outputs(results, output_path, embed_covers=args.embed_covers)
    safe_print(f"Done: {output_path}")
    safe_print(f"CSV: {output_path.with_suffix('.csv')}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="小红书笔记批量采集与分析 agent")
    parser.add_argument("--input", required=True, help="源 Excel 文件路径")
    parser.add_argument("--output", default="outputs/xhs_note_analysis.xlsx", help="输出 xlsx 路径")
    parser.add_argument("--limit", type=int, default=0, help="只处理前 N 条，0 表示全部")
    parser.add_argument("--no-crawl", action="store_true", help="不访问网页，只基于 Excel 数据做分析")
    parser.add_argument("--headless", action="store_true", help="无头浏览器模式")
    parser.add_argument("--profile", help="Playwright 持久化登录态目录")
    parser.add_argument("--browser-executable", help="Chrome/Edge 浏览器 exe 路径，默认自动发现")
    parser.add_argument("--cdp-url", help="连接已启动的真实 Chrome，例如 http://127.0.0.1:9222")
    parser.add_argument("--login-first", action="store_true", help="先停在小红书登录页，扫码后按 Enter 再开始批量采集")
    parser.add_argument("--crawl-delay", type=float, default=1.5, help="每条笔记之间的等待秒数，遇到安全验证时可调大")
    parser.add_argument("--no-stop-on-rate-limit", action="store_true", help="遇到安全验证时不中止，继续处理后续链接")
    parser.add_argument("--rate-limit-cooldown", type=int, default=0, help="遇到安全验证后的冷却秒数；为 0 时默认中止")
    parser.add_argument("--no-comment-api", action="store_true", help="关闭评论分页接口补采，只使用页面可见评论")
    parser.add_argument("--download-covers", action="store_true", help="预留：下载封面到本地")
    parser.add_argument("--embed-covers", action="store_true", help="把已下载的本地封面图片嵌入输出 Excel")
    parser.add_argument("--crawl-pgy", action="store_true", help="尝试进入蒲公英后台抓取达人报价/粉丝量，并计算 CPE")
    parser.add_argument("--pgy-delay", type=float, default=3.0, help="蒲公英后台页面间等待秒数")
    parser.add_argument("--pgy-timeout", type=int, default=30000, help="蒲公英后台页面超时毫秒")
    parser.add_argument("--pgy-safe-mode", action="store_true", help="蒲公英慢速安全模式：更长等待、更小节奏抖动")
    parser.add_argument("--pgy-max-retries", type=int, default=1, help="蒲公英单达人最大尝试次数")
    parser.add_argument("--use-llm", action="store_true", help="启用 LLM API 优化创意建议和人群圈选策略")
    parser.add_argument("--llm-api-key", help="LLM API key；也可用 LLM_API_KEY / OPENAI_API_KEY 环境变量")
    parser.add_argument("--llm-base-url", help="OpenAI-compatible base URL；默认 https://api.openai.com/v1")
    parser.add_argument("--llm-model", help="LLM 模型名；也可用 LLM_MODEL 环境变量")
    parser.add_argument("--llm-timeout", type=int, default=60, help="LLM 请求超时秒数")
    parser.add_argument("--llm-temperature", type=float, default=0.3, help="LLM 生成温度")
    return parser


if __name__ == "__main__":
    run(build_parser().parse_args())

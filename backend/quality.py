from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill


SOURCE_SHEET = "笔记明细"
QUALITY_REQUIRED_COLUMNS = [
    "文案",
    "封面",
    "达人昵称",
    "粉丝量",
    "评论区前20条",
]


def quality_report(csv_path: Path, required_columns: list[str] | None = None, limit: int = 50) -> dict[str, Any]:
    required = required_columns or QUALITY_REQUIRED_COLUMNS
    if not csv_path.exists():
        return {"rows": 0, "score": 0, "required_columns": required, "missing": {}, "retry_rows": []}
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        columns = reader.fieldnames or []
        rows = list(reader)
    total = len(rows)
    missing = {column: 0 for column in required}
    status_counts: dict[str, int] = {}
    retry_rows = []
    complete_rows = 0
    for idx, row in enumerate(rows, start=1):
        status = str(row.get("采集状态", "")).strip() or "unknown"
        status_counts[status] = status_counts.get(status, 0) + 1
        missing_columns = [column for column in required if not str(row.get(column, "")).strip()]
        for column in missing_columns:
            missing[column] += 1
        needs_retry = status != "ok" or bool(missing_columns)
        if not needs_retry:
            complete_rows += 1
        elif len(retry_rows) < limit:
            retry_rows.append(
                {
                    "index": idx,
                    "row_number": row.get("源表行号", ""),
                    "title": row.get("标题", ""),
                    "url": row.get("笔记链接", ""),
                    "status": status,
                    "missing_columns": missing_columns,
                    "error": row.get("异常信息", ""),
                }
            )
    score = round((complete_rows / total) * 100, 2) if total else 0
    return {
        "rows": total,
        "score": score,
        "complete_rows": complete_rows,
        "retry_needed": total - complete_rows,
        "required_columns": required,
        "missing": missing,
        "status_counts": status_counts,
        "columns": columns,
        "retry_rows": retry_rows,
    }


def read_result_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return reader.fieldnames or [], list(reader)


def row_missing(row: dict[str, str], required_columns: list[str]) -> bool:
    status = str(row.get("采集状态", "")).strip()
    if status and status != "ok":
        return True
    return any(not str(row.get(column, "")).strip() for column in required_columns)


def select_retry_rows(rows: list[dict[str, str]], required_columns: list[str]) -> list[dict[str, str]]:
    return [row for row in rows if row_missing(row, required_columns)]


def load_source_rows(source_path: Path) -> tuple[list[str], dict[int, list[Any]], dict[str, list[Any]]]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if SOURCE_SHEET not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {SOURCE_SHEET}")
    sheet = workbook[SOURCE_SHEET]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    by_row_number: dict[int, list[Any]] = {}
    by_url: dict[str, list[Any]] = {}
    url_index = next(
        (
            index
            for index, header in enumerate(headers)
            if str(header or "").strip().lower() in {"笔记url", "笔记链接", "链接"}
        ),
        None,
    )
    for row_number in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)]
        by_row_number[row_number] = values
        if url_index is not None:
            url = str(values[url_index] or "").strip()
            if url:
                by_url[url] = values
    return [str(header or "") for header in headers], by_row_number, by_url


def build_retry_workbook(
    source_path: Path,
    retry_rows: list[dict[str, str]],
    target_path: Path,
) -> int:
    headers, by_row_number, by_url = load_source_rows(source_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SOURCE_SHEET
    sheet.append(headers)
    added = 0
    seen: set[tuple[str, str]] = set()
    for result_row in retry_rows:
        source_values = None
        row_number_text = str(result_row.get("源表行号", "")).strip()
        if row_number_text.isdigit():
            source_values = by_row_number.get(int(row_number_text))
        if source_values is None:
            source_values = by_url.get(str(result_row.get("笔记链接", "")).strip())
        if source_values is None:
            continue
        key = (row_number_text, str(result_row.get("笔记链接", "")).strip())
        if key in seen:
            continue
        seen.add(key)
        sheet.append(source_values)
        added += 1
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target_path)
    return added


def retry_prep_report(
    source_path: Path,
    result_csv_path: Path,
    output_dir: Path,
    required_columns: list[str] | None = None,
    output_stem: str | None = None,
) -> dict[str, Any]:
    required = required_columns or QUALITY_REQUIRED_COLUMNS
    _, base_rows = read_result_csv(result_csv_path)
    retry_candidates = select_retry_rows(base_rows, required)
    stamp = datetime.now().strftime("%Y年%m月%d日%H时%M分%S秒")
    stem = safe_filename_part(output_stem or "补抓输入表")
    retry_input = unique_path(output_dir / f"{stem}_{stamp}.xlsx")
    retry_count = build_retry_workbook(source_path, retry_candidates, retry_input)
    return {
        "source": str(source_path.resolve()),
        "result_csv": str(result_csv_path.resolve()),
        "required_columns": required,
        "retry_candidates": len(retry_candidates),
        "retry_input_rows": retry_count,
        "retry_input": str(retry_input.resolve()),
    }


def safe_filename_part(value: str) -> str:
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", str(value or ""))
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"_+", "_", text).strip("._ ")
    return text[:60] or "补抓输入表"


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(2, 1000):
        candidate = path.with_name(f"{path.stem}_第{index}次{path.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"cannot find available filename for {path.name}")


def merge_rows(
    base_columns: list[str],
    base_rows: list[dict[str, str]],
    retry_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    retry_by_key: dict[str, dict[str, str]] = {}
    for row in retry_rows:
        key = str(row.get("源表行号", "")).strip() or str(row.get("笔记链接", "")).strip()
        if key:
            retry_by_key[key] = row
    merged: list[dict[str, str]] = []
    for row in base_rows:
        key = str(row.get("源表行号", "")).strip() or str(row.get("笔记链接", "")).strip()
        replacement = retry_by_key.get(key)
        if replacement and str(replacement.get("采集状态", "")).strip() == "ok":
            merged.append({column: replacement.get(column, row.get(column, "")) for column in base_columns})
        else:
            merged.append(row)
    return merged


def write_result_table(
    columns: list[str],
    rows: list[dict[str, str]],
    output_path: Path,
    embed_covers: bool = True,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path = output_path.with_suffix(".csv")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "笔记分析结果"
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "标题": 34,
        "笔记链接": 46,
        "封面": 24 if embed_covers else 40,
        "文案": 60,
        "话题": 30,
        "达人昵称": 18,
        "达人ID": 22,
        "达人链接": 42,
        "评论区前20条": 60,
        "蒲公英链接": 42,
        "蒲公英图文报价": 16,
        "蒲公英视频报价": 16,
        "图文CPE": 14,
        "视频CPE": 14,
        "创意建议": 60,
        "人群圈选策略": 60,
        "LLM状态": 26,
        "LLM模型": 22,
        "异常信息": 36,
    }
    for col_idx, header in enumerate(columns, start=1):
        letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[letter].width = widths.get(header, 16)
    if embed_covers and "封面" in columns:
        embed_cover_images(
            sheet,
            rows,
            cover_col=columns.index("封面") + 1,
            start_row=2,
            converted_dir=output_path.parent / "embedded_covers",
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_path)


def embed_cover_images(
    sheet: Any,
    rows: list[dict[str, str]],
    cover_col: int,
    start_row: int,
    converted_dir: Path,
) -> None:
    thumb_px = 120
    converted_dir.mkdir(parents=True, exist_ok=True)
    for offset, row in enumerate(rows):
        row_idx = start_row + offset
        cover_path = Path(str(row.get("封面", "") or ""))
        if not cover_path.exists() or not cover_path.is_file():
            continue
        try:
            image_path = prepare_excel_image_path(cover_path, converted_dir)
            image = ExcelImage(str(image_path))
            width = image.width or thumb_px
            height = image.height or thumb_px
            scale = min(thumb_px / width, thumb_px / height)
            image.width = int(width * scale)
            image.height = int(height * scale)
            cell = sheet.cell(row=row_idx, column=cover_col)
            cell.value = ""
            sheet.add_image(image, cell.coordinate)
            sheet.row_dimensions[row_idx].height = max(sheet.row_dimensions[row_idx].height or 15, 95)
        except Exception as exc:
            sheet.cell(row=row_idx, column=cover_col).value = f"{row.get('封面', '')} (embed failed: {exc})"


def prepare_excel_image_path(path: Path, converted_dir: Path) -> Path:
    if path.suffix.lower() in {".png", ".jpg", ".jpeg", ".bmp", ".gif"}:
        return path
    from PIL import Image

    target = converted_dir / f"{path.stem}.png"
    if target.exists() and target.stat().st_mtime >= path.stat().st_mtime:
        return target
    with Image.open(path) as img:
        if img.mode not in {"RGB", "RGBA"}:
            img = img.convert("RGBA")
        img.save(target, "PNG")
    return target


def merge_retry_results(base_csv_path: Path, retry_csv_path: Path, merged_output_path: Path) -> dict[str, Any]:
    base_columns, base_rows = read_result_csv(base_csv_path)
    retry_columns, retry_rows = read_result_csv(retry_csv_path)
    if retry_columns != base_columns:
        raise ValueError("Retry output columns differ from base result columns.")
    merged_rows = merge_rows(base_columns, base_rows, retry_rows)
    write_result_table(base_columns, merged_rows, merged_output_path)
    return {
        "merged_output": str(merged_output_path.resolve()),
        "merged_csv": str(merged_output_path.with_suffix(".csv").resolve()),
        "merged_quality": quality_report(merged_output_path.with_suffix(".csv")),
    }

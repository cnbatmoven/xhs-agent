from __future__ import annotations

import argparse
import csv
import tempfile
import sys
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from xhs_note_agent import run

OUTPUT_DIR = BASE_DIR / "outputs"
SOURCE_SHEET = "\u7b14\u8bb0\u660e\u7ec6"
RESULT_SHEET = "\u7b14\u8bb0\u5206\u6790\u7ed3\u679c"
ROW_COL = "\u6e90\u8868\u884c\u53f7"
URL_COL = "\u7b14\u8bb0\u94fe\u63a5"
STATUS_COL = "\u91c7\u96c6\u72b6\u6001"
DEFAULT_REQUIRED = [
    "\u6587\u6848",
    "\u5c01\u9762",
    "\u8fbe\u4eba\u6635\u79f0",
    "\u7c89\u4e1d\u91cf",
    "\u8bc4\u8bba\u533a\u524d20\u6761",
]


def read_result_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return reader.fieldnames or [], list(reader)


def result_csv_path(path: Path) -> Path:
    if path.suffix.lower() == ".csv":
        return path
    return path.with_suffix(".csv")


def row_missing(row: dict[str, str], required_columns: list[str]) -> bool:
    status = str(row.get(STATUS_COL, "")).strip()
    if status and status != "ok":
        return True
    return any(not str(row.get(column, "")).strip() for column in required_columns)


def select_retry_rows(rows: list[dict[str, str]], required_columns: list[str]) -> list[dict[str, str]]:
    return [row for row in rows if row_missing(row, required_columns)]


def load_source_rows(source_path: Path) -> tuple[list[str], dict[int, list[Any]], dict[str, list[Any]]]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if SOURCE_SHEET not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {SOURCE_SHEET}")
    sheet = workbook[SOURCE_SHEET]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    by_row_number: dict[int, list[Any]] = {}
    by_url: dict[str, list[Any]] = {}
    url_index = next(
        (
            index
            for index, header in enumerate(headers)
            if str(header or "").strip().lower() in {"\u7b14\u8bb0url", "\u7b14\u8bb0\u94fe\u63a5", "\u94fe\u63a5"}
        ),
        None,
    )
    for row_number in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)]
        by_row_number[row_number] = values
        if url_index is not None:
            url = str(values[url_index] or "").strip()
            if url:
                by_url[url] = values
    return [str(header or "") for header in headers], by_row_number, by_url


def build_retry_workbook(
    source_path: Path,
    retry_rows: list[dict[str, str]],
    target_path: Path,
) -> int:
    headers, by_row_number, by_url = load_source_rows(source_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SOURCE_SHEET
    sheet.append(headers)
    added = 0
    seen: set[tuple[str, str]] = set()
    for result_row in retry_rows:
        source_values = None
        row_number_text = str(result_row.get(ROW_COL, "")).strip()
        if row_number_text.isdigit():
            source_values = by_row_number.get(int(row_number_text))
        if source_values is None:
            source_values = by_url.get(str(result_row.get(URL_COL, "")).strip())
        if source_values is None:
            continue
        key = (row_number_text, str(result_row.get(URL_COL, "")).strip())
        if key in seen:
            continue
        seen.add(key)
        sheet.append(source_values)
        added += 1
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target_path)
    return added


def run_retry(source_path: Path, retry_input: Path, retry_output: Path, args: argparse.Namespace) -> None:
    payload = {
        "input": str(retry_input),
        "output": str(retry_output),
        "limit": 0,
        "no_crawl": False,
        "headless": True,
        "profile": None,
        "browser_executable": None,
        "cdp_url": args.cdp_url,
        "login_first": False,
        "crawl_delay": args.crawl_delay,
        "no_stop_on_rate_limit": False,
        "rate_limit_cooldown": 0,
        "no_comment_api": False,
        "download_covers": args.download_covers,
        "embed_covers": args.embed_covers,
        "crawl_pgy": args.crawl_pgy,
        "pgy_delay": args.pgy_delay,
        "pgy_timeout": 30000,
        "pgy_safe_mode": args.pgy_safe_mode,
        "pgy_max_retries": args.pgy_max_retries,
        "use_llm": args.use_llm,
        "llm_api_key": None,
        "llm_base_url": args.llm_base_url,
        "llm_model": args.llm_model,
        "llm_timeout": 60,
        "llm_temperature": 0.3,
    }
    run(SimpleNamespace(**payload))


def merge_rows(
    base_columns: list[str],
    base_rows: list[dict[str, str]],
    retry_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    retry_by_key: dict[str, dict[str, str]] = {}
    for row in retry_rows:
        key = str(row.get(ROW_COL, "")).strip() or str(row.get(URL_COL, "")).strip()
        if key:
            retry_by_key[key] = row
    merged: list[dict[str, str]] = []
    for row in base_rows:
        key = str(row.get(ROW_COL, "")).strip() or str(row.get(URL_COL, "")).strip()
        replacement = retry_by_key.get(key)
        if replacement and str(replacement.get(STATUS_COL, "")).strip() == "ok":
            merged.append({column: replacement.get(column, row.get(column, "")) for column in base_columns})
        else:
            merged.append(row)
    return merged


def write_outputs(columns: list[str], rows: list[dict[str, str]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path = output_path.with_suffix(".csv")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = RESULT_SHEET
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for idx, column in enumerate(columns, start=1):
        letter = sheet.cell(1, idx).column_letter
        sheet.column_dimensions[letter].width = 42 if column in {URL_COL, "\u6587\u6848"} else 16
    workbook.save(output_path)


def summarize(rows: list[dict[str, str]], required_columns: list[str]) -> dict[str, int]:
    return {
        "rows": len(rows),
        "status_ok": sum(1 for row in rows if str(row.get(STATUS_COL, "")).strip() == "ok"),
        "retry_needed": sum(1 for row in rows if row_missing(row, required_columns)),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Retry failed or incomplete XHS result rows and merge them back.")
    parser.add_argument("--source", required=True, help="Original workbook with 笔记明细 sheet.")
    parser.add_argument("--result", required=True, help="Previous result .xlsx or .csv.")
    parser.add_argument("--output", default="")
    parser.add_argument("--required", default=",".join(DEFAULT_REQUIRED))
    parser.add_argument("--cdp-url", default="http://127.0.0.1:9222")
    parser.add_argument("--crawl-delay", type=float, default=10.0)
    parser.add_argument("--download-covers", action="store_true")
    parser.add_argument("--embed-covers", action="store_true")
    parser.add_argument("--crawl-pgy", action="store_true")
    parser.add_argument("--pgy-delay", type=float, default=12.0)
    parser.add_argument("--pgy-safe-mode", action="store_true")
    parser.add_argument("--pgy-max-retries", type=int, default=2)
    parser.add_argument("--use-llm", action="store_true")
    parser.add_argument("--llm-base-url", default=None)
    parser.add_argument("--llm-model", default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    source_path = Path(args.source).expanduser().resolve()
    result_path = result_csv_path(Path(args.result).expanduser().resolve())
    required_columns = [item.strip() for item in args.required.split(",") if item.strip()]
    columns, base_rows = read_result_csv(result_path)
    retry_candidates = select_retry_rows(base_rows, required_columns)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else OUTPUT_DIR / f"{Path(args.result).stem}_retry_merged_{stamp}.xlsx"
    )
    retry_input = OUTPUT_DIR / f"retry_input_{stamp}.xlsx"
    retry_output = OUTPUT_DIR / f"retry_output_{stamp}.xlsx"
    retry_count = build_retry_workbook(source_path, retry_candidates, retry_input)

    report: dict[str, Any] = {
        "source": str(source_path),
        "result": str(result_path),
        "output": str(output_path),
        "required_columns": required_columns,
        "base_summary": summarize(base_rows, required_columns),
        "retry_candidates": len(retry_candidates),
        "retry_input_rows": retry_count,
        "retry_input": str(retry_input),
    }
    if args.dry_run or retry_count == 0:
        print(report)
        return 0

    run_retry(source_path, retry_input, retry_output, args)
    retry_columns, new_rows = read_result_csv(retry_output.with_suffix(".csv"))
    if retry_columns != columns:
        raise ValueError("Retry output columns differ from base result columns.")
    merged_rows = merge_rows(columns, base_rows, new_rows)
    write_outputs(columns, merged_rows, output_path)
    report.update(
        {
            "retry_output": str(retry_output),
            "retry_summary": summarize(new_rows, required_columns),
            "merged_summary": summarize(merged_rows, required_columns),
            "merged_csv": str(output_path.with_suffix(".csv")),
        }
    )
    print(report)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
